import openpyxl


wb = openpyxl.load_workbook('scan_report.xlsx')
sheet = wb['scanner-sample']
no_detection = ["No Detection", None, "NA", "Blank", "", " "]
nd_list = []


for row in range(2, sheet.max_row + 1):
    if sheet.cell(row=row, column=1).value == None:
        break
    cell_G = sheet.cell(row=row, column=7).value
    cell_J = sheet.cell(row=row, column=10).value
    if (cell_G in no_detection) and (cell_J in no_detection):
        nd_list.append(row)
        # print(cell_G)

# print(nd_list)

workbook = openpyxl.load_workbook("scan_report.xlsx")
sheet = workbook["scanner-sample"]
cell = sheet["N1"]
cell.value = "ND"

for i in nd_list:
    sheet[f"N{i}"].value = "No Detection"
    pass

workbook.save("scan_report.xlsx")

print("created new No Detection column")

pe_list = []
nonpe_list = []
for row in range(2, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=14).value
    if cell_value == "No Detection":
        if sheet.cell(row=row, column=5).value == "exe":
            pe_list.append(sheet.cell(row=row, column=1).value)
        else:
            nonpe_list.append(sheet.cell(row=row, column=1).value)
        pass

# print(pe_list, nonpe_list)

workbook = openpyxl.load_workbook("scan_report.xlsx")
sheet = workbook["undetected"]
n = 2
for i in pe_list:
    sheet[f"A{n}"].value = i
    n += 1
    pass
n = 2
for i in nonpe_list:
    sheet[f"B{n}"].value = i
    n += 1
    pass

workbook.save("scan_report.xlsx")

print("PE and NON-PE saved to undetected sheet")

