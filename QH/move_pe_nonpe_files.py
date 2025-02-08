import openpyxl
import os

def move_PE_NONPE_files():
    wb = openpyxl.load_workbook('scan_report.xlsx')
    sheet = wb['undetected']
    PE_values = []
    NON_PE_values = []

    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value != None:
            PE_values.append(cell_value)

    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=2).value
        if cell_value != None:
            NON_PE_values.append(cell_value)

    print("PE_values:", len(PE_values))
    print("NON_PE_values:", len(NON_PE_values))
    if not os.path.exists("PE"):
        os.mkdir("PE")
    if not os.path.exists("NON_PE"):
        os.mkdir("NON_PE")
    
    for i in PE_values:
        try:
            os.rename(f"./sample/{i}", f"./PE/{i}")
        except Exception as e:
            print(e)

    for i in NON_PE_values:
        try:
            os.rename(f"./sample/{i}", f"./NON_PE/{i}")
        except Exception as e:
            print(e)

if __name__=='__main__':
    # create xlsx file name 'scan_report.xlsx' inside sheet name 'undetected'
    # place PE files in first column & NON-PE files to second column under sheet named 'undetected'
    # start from second row, first row is title
    move_PE_NONPE_files()

print("\n\nDone...")

